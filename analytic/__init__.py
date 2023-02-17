import xlrd
import openpyxl
import configparser
from datetime import datetime
from openpyxl.writer.excel import save_workbook

config = configparser.ConfigParser()
config.read("config.ini")


def data_extraction():
    read_book = xlrd.open_workbook(f"{config['CONFIG']['save']}.xlsx")
    number_sheets = (len(read_book.sheet_names()) - 1)  # Единица вычитается, так как отсчет идет с 0
    read_sheet_penultimate = read_book.sheet_by_index(number_sheets - 1)
    read_sheet_last = read_book.sheet_by_index(number_sheets)

    if len(read_book.sheet_names()) >= 1:
        penultimate_list = []
        last_list = []
        quantity_penultimate = read_sheet_penultimate.nrows
        quantity_last = read_sheet_last.nrows
        for iter_line in (range(1, max(quantity_penultimate, quantity_last))):
            penultimate_list_temp = []
            last_list_temp = []
            for iter_column in range(0, 7):

                try:
                    penultimate_list_temp.append(read_sheet_penultimate.cell_value(iter_line, iter_column))
                except IndexError:
                    pass

                try:
                    last_list_temp.append(read_sheet_last.cell_value(iter_line, iter_column))
                except IndexError:
                    pass

            penultimate_list.append(penultimate_list_temp)
            last_list.append(last_list_temp)

        return penultimate_list, last_list
    else:
        print("Для анализа данных необходимо минимум 2 листа данных!!!")


def comparison():
    penultimate_, last_ = data_extraction()
    analytic_list = []
    for val_penultimate in penultimate_:
        for val_last in last_:
            try:
                if val_penultimate[2] == val_last[2]:
                    if val_penultimate[3] != val_last[3] or val_penultimate[4] != val_last[4] or val_penultimate[6] != \
                            val_last[6]:
                        analytic_list.append(val_last)
            except IndexError:
                pass
    return analytic_list


def save_analytic():
    sheet_name = f"{config['CONFIG']['save_analytics']}.xlsx"
    enter_datatime = f"{datetime.now().day}-" \
                     f"{datetime.now().month}-" \
                     f"{datetime.now().year}_" \
                     f"{datetime.now().hour}-" \
                     f"{datetime.now().minute}"
    columns = ['ССЫЛКА', 'НАИМЕНОВАНИЕ', 'АРТИКУЛ', 'ПОСЛЕДНЕЕ ПРЕДЛОЖЕНИЕ', 'ПОСЛЕДНЯЯ ЦЕНА ПРОДАЖИ', 'ДАТА ВЫПУСКА',
               'КОЛИЧЕСТВО ЖЕЛАЮЩИХ', 'ЦЕНА ДОСТАВКИ']

    try:
        wb = openpyxl.load_workbook(sheet_name)
    except:
        wb = openpyxl.Workbook()

        for sheet_names in wb.sheetnames:
            sheet = wb[f"{sheet_names}"]
            wb.remove(sheet)

    ws = wb.create_sheet(f"{config['CONFIG']['search']}_{enter_datatime}")

    for i, value in enumerate(columns, 1):
        ws.cell(row=1, column=i).value = value

    for i, row in enumerate(comparison(), 2):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j).value = value

    save_workbook(wb, sheet_name)
