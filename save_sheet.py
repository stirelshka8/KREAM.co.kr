import openpyxl
import configparser
from datetime import datetime
from updater import _extract_data
from openpyxl.writer.excel import save_workbook

config = configparser.ConfigParser()
config.read("config.ini")

enter_datatime = f"{datetime.now().day}-" \
                 f"{datetime.now().month}-" \
                 f"{datetime.now().year}_" \
                 f"{datetime.now().hour}-" \
                 f"{datetime.now().minute}"

columns = ['ССЫЛКА', 'НАИМЕНОВАНИЕ', 'АРТИКУЛ', 'ПОСЛЕДНЕЕ ПРЕДЛОЖЕНИЕ', 'ПОСЛЕДНЯЯ ЦЕНА ПРОДАЖИ', 'ДАТА ВЫПУСКА',
           'КОЛИЧЕСТВО ЖЕЛАЮЩИХ']

FILE_NAME = f"{config['CONFIG']['save']}.xlsx"
SHEET_NAME = f"{config['CONFIG']['search']}_{enter_datatime}"


def save():
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
    except:
        wb = openpyxl.Workbook()

        for sheet_name in wb.sheetnames:
            sheet = wb[f"{sheet_name}"]
            wb.remove(sheet)

    ws = wb.create_sheet(SHEET_NAME)

    for i, value in enumerate(columns, 1):
        ws.cell(row=1, column=i).value = value

    for i, row in enumerate(_extract_data(), 2):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j).value = value

    save_workbook(wb, f"{config['CONFIG']['save']}.xlsx")

